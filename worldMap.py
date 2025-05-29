import os
import subprocess
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment
from models.land import Land

def displayMap(selected_options):
    # Create a new workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Set the border style for all cells
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Set column width and row height for all cells
    column_width = 20
    row_height = 80
    for i in range(8, 18):  # Adjusted range to match a 7x10 grid
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = column_width
    for i in range(8, 15):  # Adjusted range to match a 7x10 grid
        sheet.row_dimensions[i].height = row_height

    # Create a 2D array to store the locations of capitals and empty cells
    # Create individual Land objects and assign them to specific indices in the 2D array
    world_map = [
        [Land("Land 1"), Land("Land 2"), Land("Land 3"), Land("Land 4"), Land("Land 5"), Land("Land 6"), Land("Land 7"), Land("Land 8"), Land("Land 9"), Land("Land 10")],
        [Land("Land 11"), Land("Land 12"), Land("Land 13"), Land("Land 14"), Land("Land 15"), Land("Land 16"), Land("Land 17"), Land("Land 18"), Land("Land 19"), Land("Land 20")],
        [Land("Land 21"), Land("Land 22"), Land("Land 23"), Land("Land 24"), Land("Land 25"), Land("Land 26"), Land("Land 27"), Land("Land 28"), Land("Land 29"), Land("Land 30")],
        [Land("Land 31"), Land("Land 32"), Land("Land 33"), Land("Land 34"), Land("Land 35"), Land("Land 36"), Land("Land 37"), Land("Land 38"), Land("Land 39"), Land("Land 40")],
        [Land("Land 41"), Land("Land 42"), Land("Land 43"), Land("Land 44"), Land("Land 45"), Land("Land 46"), Land("Land 47"), Land("Land 48"), Land("Land 49"), Land("Land 50")],
        [Land("Land 51"), Land("Land 52"), Land("Land 53"), Land("Land 54"), Land("Land 55"), Land("Land 56"), Land("Land 57"), Land("Land 58"), Land("Land 59"), Land("Land 60")],
        [Land("Land 61"), Land("Land 62"), Land("Land 63"), Land("Land 64"), Land("Land 65"), Land("Land 66"), Land("Land 67"), Land("Land 68"), Land("Land 69"), Land("Land 70")]
    ]

    # Loop through the list of kingdoms and set the background color and value of each cell
    previous_choices = []
    
    for kingdom_key in selected_options:
        kingdom = selected_options[kingdom_key]
        capital = kingdom.capital
        capital_name = capital.name
        kingdom_color = kingdom.color
        capital_location = input(f"{kingdom.name}, where would you like to locate your capital? (must be greater than or equal to 1x1 and less than or equal to 7x10): ")
        while True:
            try:
                capital_row, capital_column = map(int, capital_location.split())

                capital_column += 7
                capital_row += 7

                if capital_column < 8 or capital_column > 17 or capital_row < 8 or capital_row > 14:
                    raise ValueError("you are out of bounds. Remember it must be greater than or equal to 1x1 and less than or equal to 7x10 ")
                
                if [capital_row, capital_column] in previous_choices:
                    raise ValueError("That choice has already been selected.")

                previous_choices.append([capital_row, capital_column])

                # Clear previous values and formatting of the selected cell
                cell = sheet.cell(row=capital_row, column=capital_column)
                cell.value = None
                cell.fill = PatternFill(start_color=None, end_color=None, fill_type=None)
                cell.border = thin_border

                # Set the value and formatting of the new selected cell
                cell.value = capital_name + "*\n" + ", ".join([character.name for land in kingdom.lands for character in land.charactersPresent])
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                cell.fill = PatternFill(start_color=kingdom_color, end_color=kingdom_color, fill_type='solid')
                cell.border = thin_border

                # Update the Land object's position in the world_map array
                capital.row, capital.column = capital_row - 8, capital_column - 8
                world_map[capital.row][capital.column] = capital
            except ValueError as e:
                print(e)
                capital_location = input("Please try again: ")
                continue
        
            break
    # Apply formatting to all cells in the world map
    for row in range(8, 15):
        for column in range(8, 18):
            land = world_map[row - 8][column - 8]
            cell = sheet.cell(row=row, column=column)

            if land.is_capital:
            # Skip formatting for capital cells
                continue
            else:
                # Set formatting for land cells
                cell.value = land.name
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                cell.border = thin_border

    # Save the workbook
    workbook.save('World Map.xlsx')

    # Open the saved file using Excel
    if os.name == 'nt':  # Windows
        os.startfile('World Map.xlsx')
    elif os.name == 'posix':  # macOS or Linux
        subprocess.call(('open', 'World Map.xlsx'))

    # input(world_map)
    return world_map
