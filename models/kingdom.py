import xlwings as xw
from openpyxl.utils import get_column_letter
from termcolor import colored
from config.game_constants import MARKET_PRICES, PURCHASE_WAIT_TIMES
from models.battalion import Battalion
from models.land import Land
from utils.display_utils import clear_screen, center

class Kingdom:
    def __init__(self, playerNumber, name, capitalName, ruler, advisor, warrior, color):
        self.playerNumber = playerNumber
        self.name = name
        self.ruler = Battalion(5, ruler, 3, True)
        self.advisor = Battalion(5, advisor, 3)
        self.capital = Land(capitalName, name, 3, 12, [self.ruler, self.advisor], True)
        self.lands = [self.capital]
        self.warrior = warrior
        self.purchase_delay = {}
        self.purchaseWaitTime = PURCHASE_WAIT_TIMES
        self.color = color
        self.gold = 21
        self.defeated = False
        self.updateVariables()

    def updateVariables(self):
        productionPower = 0
        numCitiesCounter = 0
        for land in self.lands:
            productionPower += land.numPeasants
            numCitiesCounter += 1
        self.production = productionPower
        self.numCities = numCitiesCounter

        if self.purchase_delay:
            for countdown in range(max(self.purchase_delay.keys()), -1, -1):
                if countdown in self.purchase_delay:
                    items = self.purchase_delay[countdown]
                    if countdown == 1:
                        self.addItem(items)
                        del self.purchase_delay[countdown]
                    else:
                        self.purchase_delay[countdown] = self.purchase_delay.get(countdown - 1, [])
                        if not self.purchase_delay[countdown - 1]:
                            del self.purchase_delay[countdown - 1]

    def updateExcel(self):
        wb = xw.Book('World Map.xlsx')
        sht1 = wb.sheets['Sheet1']

        for land in self.lands:
            row = land.row + 8
            column = land.column + 8

            character_info = []
            for char in land.charactersPresent:
                if char.name == "Battalion":
                    character_info.append(str(char.level))
                else:
                    character_info.append(char.name)

            land_info = land.name + '\n' + ', '.join(character_info)
            cell = sht1.range((row, column))
            cell.value = land_info

        wb.save()

    def canMoveStatus(self, character, turn):
        if(character.lastTurnTheyMoved < turn):
            return colored(str(character.level), 'green')
        else:
            return colored(str(character.level),'red')

    def printCharacters(self, charactersPresent, turn):
        result = ""
        for character in charactersPresent:
            if character == self.ruler or character == self.advisor or character.name == "C":
                result += character.name + "<"
            result += "(" + self.canMoveStatus(character, turn) + ") "
        return result
    
    def printPurchaseDelay(self, city_index):
        if city_index not in range(len(self.lands)):
            return ""
        if self.purchase_delay:
            output = ""
            for key in self.purchase_delay:
                if isinstance(self.purchase_delay[key][0], list):
                    for arr in self.purchase_delay[key]:
                        if arr[0] == city_index:
                            output += f"{'' if arr[1] == 1 else arr[1]}{arr[2].upper()}({self.purchaseWaitTime[arr[2]] - key}/{self.purchaseWaitTime[arr[2]]}) "
                else:
                    if self.purchase_delay[key][0] == city_index:
                        output += f"{'' if self.purchase_delay[key][1] == 1 else self.purchase_delay[key][1]}{self.purchase_delay[key][2]}{self.purchaseWaitTime[self.purchase_delay[key][2]] - key}/{self.purchaseWaitTime[self.purchase_delay[key][2]]} "
            return output
        else:
            return ""

    def printCities(self, turn):
        for i in range(len(self.lands)):
            if self.capital == self.lands[i]:
                capitalStar = "*"
            else:
                capitalStar = ""
            print("\n\t" + str(i+1) + " - " + capitalStar + self.lands[i].name + capitalStar + "\n\t\tPeasants: " + str(self.lands[i].numPeasants) + "/" + str(self.lands[i].housing) + "    Idle: " + str(self.lands[i].numAvailablePeasants) + "/" + str(self.lands[i].numPeasants))
            print("\t\tProgress: " + self.printPurchaseDelay(i))
            print("\t\tBattalions: " + self.printCharacters(self.lands[i].charactersPresent, turn))

    def buy(self, item):
        while True:
            if item == "x":
                return
            if item == "shibboleth":
                self.addItem(input("How much money do you want to add? "))
            request = item.split()
            if len(request) != 3:
                item = input("Invalid input format\nmust be 3 elements long with spaces in between:\n")
                continue

            try:
                city_index = int(request[0]) - 1
                quantity = int(request[1])
            except ValueError:
                item = input("Invalid input format\nmust be: 'number, number, letter':\n")
                continue

            item_type = request[2].lower()

            if item_type not in MARKET_PRICES:
                item = input("Invalid item type\nmust be one of: 'p', 'g', 'b', 'm', 't', 'c', 'h', 'hp' or press x to exit:\n")
                continue

            if self.gold < quantity * MARKET_PRICES[item_type]:
                item = input("Insufficient gold to purchase the requested quantity of this item:\n")
                continue

            if city_index >= len(self.lands) or city_index < 0:
                item = input("You selected a kingdom that does not exist. \nPlease retry or exit with x:\n")
                continue

            if self.lands[city_index].numAvailablePeasants < quantity:
                item = input("Insufficient peasants to build the requested quantity of this item:\n")
                continue

            break

        self.gold -= quantity * MARKET_PRICES[item_type]

        if item_type == "hp" or item_type == "c" or item_type == "t":
            self.addItem([city_index, quantity, item_type])
            return
        else:
            self.lands[city_index].numAvailablePeasants -= quantity
            self.purchase_delay.setdefault(1, []).append([city_index, quantity, item_type])

    def addItem(self, request):
        if type(request) == str:
            request = request.split()
            request[0] = int(request[0]) - 1 
            request[1] = int(request[1])
        if type(request) == list:
            if type(request[0]) == list:
                for array in request:
                    self.addItem(array)
            else:
                city_index = request[0]
                quantity = request[1]
                item_type = request[2]

                if item_type == "p":
                    self.lands[city_index].numPeasants += quantity
                    self.lands[city_index].numAvailablePeasants += quantity
                elif item_type == "g":
                    self.gold += quantity
                elif item_type == "b":
                    for _ in range(quantity):
                        self.lands[city_index].charactersPresent.append(Battalion())
                elif item_type == "m":
                    # Not yet implemented
                    print("NOT YET IMPLEMENTED")
                elif item_type == "t":
                    # Not yet implemented
                    print("NOT YET IMPLEMENTED")
                elif item_type == "c":
                    for _ in range(quantity):
                        self.lands[city_index].charactersPresent.append(Battalion(3, "C"))
                    return
                elif item_type == "h":
                    self.lands[city_index].housing += quantity * 4
                elif item_type == "hp":
                    # Not yet implemented
                    print("NOT YET IMPLEMENTED")
                else:
                    print(f"There is an error in the request to add an item. {request} is not formatted correctly! Sucks to suck!")

                self.lands[city_index].numAvailablePeasants += quantity

    def battleScreen(self, character_land, characterObjectArray, target_land):
        clear_screen()
        center("-","-")
        center("=","=")
        center(f"   {character_land.owner} Has Attacked {target_land.owner}!!!   ","-")
        center("=","=")
        center("-","-")
        center("-"*len(f"|{character_land.owner}'s army|")," ")
        center(f"|{character_land.owner}'s army|"," ")
        center("-"*len(f"|{character_land.owner}'s army|")," ")
        center(" "," ")
        attackerString = ''
        defenderString = ''
        
        for i in range(len(characterObjectArray)):
            if characterObjectArray[i].name != "Battalion":
                attackerString += characterObjectArray[i].name + "<"
                attackerString += "(" + str(characterObjectArray[i].level) + ")"
                attackerString += " " + characterObjectArray[i].display_health()
                attackerString += " "
            else:
                attackerString += "(" + str(characterObjectArray[i].level) + ")"

        for i in range(len(target_land.charactersPresent)):
            if target_land.charactersPresent[i].name != "Battalion":
                defenderString += target_land.charactersPresent[i].name + "<"
                defenderString += "(" + str(target_land.charactersPresent[i].level) + ")"
                defenderString += " " + target_land.charactersPresent[i].display_health()
                defenderString += " "
            else:
                defenderString += "(" + str(target_land.charactersPresent[i].level) + ")"

        center(attackerString," ")
        center("-","-")
        center("-"*len(f"|{target_land.owner}'s army|")," ")
        center(f"|{target_land.owner}'s army|"," ")
        center("-"*len(f"|{target_land.owner}'s army|")," ")
        center(" "," ")
        center(defenderString," ")
        center("","=")

        numbers = [0.14, 0.33, 0.60, 1, 1.66, 3, 7]
        percentage = len(characterObjectArray) / len(target_land.charactersPresent)
        closest_number = min(numbers, key=lambda x: abs(x - percentage))
        output_strings = {
            0.14: "MAX: 1 Attacker vs. 7 Defenders",
            0.33: "MAX: 2 Attackers vs. 6 Defenders",
            0.60: "MAX: 3 Attackers vs. 5 Defenders",
            1: "MAX: 4 Attackers vs. 4 Defenders",
            1.66: "MAX: 5 Attackers vs. 3 Defenders",
            3: "MAX: 6 Attackers vs. 2 Defenders",
            7: "MAX: 7 Attackers vs 1 Defender"
        }
        center(f"|| {output_strings[closest_number]} ||","-")
        center("","=")

        print("    1: Attacker Retreat    3: Input Casualties    5: Something Else")
        print("    2: Defender Retreat    4: Use Artifacts       6: And Something Else")
        return

    def battle(self, character_land, characterObjectArray, target_land, world_map, kingdoms):
        hasAttackerMoved = False
        while True:
            self.battleScreen(character_land, characterObjectArray, target_land)
            option = input("Select an Option: ")
            
            if option == "1":
                return hasAttackerMoved
            elif option == "2":
                from config.game_constants import MOVEMENT_DIRECTIONS, VALID_MOVE_FORMATS
                retreatDirection = input("which direction do you want to retreat? ")
                
                while True:
                    while retreatDirection not in MOVEMENT_DIRECTIONS:
                        self.battleScreen(character_land, characterObjectArray, target_land)
                        retreatDirection = input("invalid direction. Please type direction or u, d, l, r as an abbreviation ")
                
                    move_row, move_col = MOVEMENT_DIRECTIONS[retreatDirection]
                    target_row = target_land.row + move_row + 8
                    target_col = target_land.column + move_col + 8

                    if not (8 <= target_row < (len(world_map)+8) and 8 <= target_col < (len(world_map[0])+8)):
                        self.battleScreen(character_land, characterObjectArray, target_land)
                        retreatDirection = input("Cannot move in the specified direction because the target location is outside the map boundaries. \nPlease retype direction:\n")
                        continue

                    new_target_land = world_map[target_row-8][target_col-8]
                    new_target_land.row = target_row-8
                    new_target_land.column = target_col-8

                    if new_target_land.owner != target_land.owner:
                        self.battleScreen(character_land, characterObjectArray, target_land)
                        retreatDirection = input("You are trying to retreat to a land that you don't own. Please choose another land.\nPlease retype direction:\n")
                        continue

                    targetLandIndexForRetreatingKingdom = kingdoms[new_target_land.owner].lands.index(target_land)
                    targetLandIndexForConqueringKingdom = len(kingdoms[character_land.owner].lands)

                    self.move(target_land, target_land.charactersPresent, new_target_land)
                    kingdoms[new_target_land.owner].remove_land(target_land)
                    self.move(character_land, characterObjectArray, target_land, True)

                    purchase_delay_change = None
                    turns_til_completion = None
                    indexesToDeleteFromNewTargetLandsPurchaseDelay = []
                    
                    for key, value in list(kingdoms[new_target_land.owner].purchase_delay.items()):
                        for i in range(len(value)):
                            if value[i][0] == targetLandIndexForRetreatingKingdom:
                                purchase_delay_change = value[i]
                                turns_til_completion = key if purchase_delay_change[2] == "b" else key - 1
                                indexesToDeleteFromNewTargetLandsPurchaseDelay.append(i)

                            if purchase_delay_change != None:
                                newEntry = [targetLandIndexForConqueringKingdom, purchase_delay_change[1], purchase_delay_change[2]]

                                if turns_til_completion or purchase_delay_change[2] == 'b':
                                    self.purchase_delay.setdefault(1, []).append(newEntry)
                                elif turns_til_completion == 0 and purchase_delay_change[2] != "b":
                                    self.addItem(newEntry)
                                
                                purchase_delay_change = None
                        
                        for index in indexesToDeleteFromNewTargetLandsPurchaseDelay:
                            value.pop(index)

                        kingdoms[character_land.owner].updateExcel()
                        kingdoms[new_target_land.owner].updateExcel()

                        hasAttackerMoved = True
                        return hasAttackerMoved

            elif option == "3":
                abbreviations = {'g': "Ganondorf", 'z': "Zelda", 'pt': "Palutena", 'n': "Ness", 'b': "Bowser", 'bj': "Bowser Jr.", 'l': "Lucina", 'i': "Ike", 'p': "Peach", 'r': "Rosalina", 'zs': "Zero Suit Samus", 's': "Samus", 'lc': "Lucario", 'gr': "Greninja", 'kd': "King Dedede", 'k': "Kirby"}

                allCharactersString = input(f"please select all of {character_land.owner}'s casualties:\n")
                charactersStringArray = allCharactersString.split(" ")

                for i in range(len(charactersStringArray)):
                    charactersStringArray[i] = charactersStringArray[i].lstrip(" ,\n")
                    charactersStringArray[i] = charactersStringArray[i].rstrip(" ,\n")

                for i in range(len(charactersStringArray)):
                    if charactersStringArray[i] in abbreviations:
                        charactersStringArray[i] = abbreviations[charactersStringArray[i]]

                for characterName in charactersStringArray:
                    if allCharactersString == "x" or allCharactersString == "":
                        break 
                    
                    hasAttackerMoved = True
    
                    if not any(char.name == characterName or (char.name == "Battalion" and str(char.level) == characterName) for char in character_land.charactersPresent):
                        move_input = input(f'Invalid character. "{characterName}" is not present in the attacking city. \nPlease enter request separated by a comma and space:\n')
                        continue
                    else:
                        for i in range(len(character_land.charactersPresent)):
                            char = character_land.charactersPresent[i]
                            if char.name == characterName or (char.name == "Battalion" and str(char.level) == characterName):
                                if (char.decrease_health()):
                                    character_land.charactersPresent.pop(i)
                                break

                allCharactersString = input(f"please select all of {target_land.owner}'s casualties:\n")
                charactersStringArray = allCharactersString.split(" ")

                for i in range(len(charactersStringArray)):
                    charactersStringArray[i] = charactersStringArray[i].lstrip(" ,\n")
                    charactersStringArray[i] = charactersStringArray[i].rstrip(" ,\n")

                for i in range(len(charactersStringArray)):
                    if charactersStringArray[i] in abbreviations:
                        charactersStringArray[i] = abbreviations[charactersStringArray[i]]

                for characterName in charactersStringArray:
                    if allCharactersString == "x" or allCharactersString == "":
                        break 
                    
                    hasAttackerMoved = True
    
                    if not any(char.name == characterName or (char.name == "Battalion" and str(char.level) == characterName) for char in character_land.charactersPresent):
                        move_input = input(f'Invalid character. "{characterName}" is not present in the attacking city. \nPlease enter request separated by a comma and space:\n')
                        continue
                    else:
                        for i in range(len(target_land.charactersPresent)):
                            char = target_land.charactersPresent[i]
                            if char.name == characterName or (char.name == "Battalion" and str(char.level) == characterName):
                                if (char.decrease_health()):
                                    target_land.charactersPresent.pop(i)
                                break

                kingdoms[character_land.owner].updateExcel()
                kingdoms[target_land.owner].updateExcel()
                continue

            elif option == "4":
                pass
            hasAttackerMoved = True

        return hasAttackerMoved

    def move(self, starting_land, characters_to_move, target_land, colonize=False):
        wb = xw.Book('World Map.xlsx')
        sht1 = wb.sheets['Sheet1']

        for character in characters_to_move.copy():
            starting_land.charactersPresent.remove(character)
            target_land.charactersPresent.append(character)

        target_column_letter = get_column_letter(target_land.column + 8)
        target_cell = sht1.range(target_column_letter + str(target_land.row + 8))

        if colonize:
            target_land.owner = self.name
            target_cell.color = self.color
            self.add_land(target_land)

        self.updateExcel()

    def moveRequest(self, world_map, move_input, turn, kingdoms):
        from config.game_constants import MOVEMENT_DIRECTIONS, VALID_MOVE_FORMATS
        
        while True:
            error = False

            if move_input == "x":
                return

            move_data = move_input.split()
            if len(move_data) != 3:
                move_input = input("Invalid input format. \nPlease provide three elements separated by spaces with the format [location] [character] [direction] or press x to exit:\n")
                continue

            city_index = int(move_data[0]) - 1
            characterNameInput = move_data[1]
            direction = move_data[2].lower()

            if city_index < 0 or city_index >= len(self.lands):
                move_input = input("Invalid city index. \nPlease select a valid city with the format [location] [character] [direction] or press x to exit:\n")
                continue

            character_land = self.lands[city_index]

            if direction not in VALID_MOVE_FORMATS:
                move_input = input("Invalid direction. \nPlease choose one of the following: 'up', 'right', 'down', 'left', or use their abbreviations 'u', 'r', 'd', 'l'. \nStart from the beginning with the format [location] [character] [direction] or press x to exit:\n")
                continue

            from config.game_constants import CHARACTER_ABBREVIATIONS
            
            if characterNameInput == "M":
                allCharactersString = input(f"please select all characters in {character_land.name} you wish to move:\n")
                charactersStringArray = allCharactersString.split(" ")
                for i in range(len(charactersStringArray)):
                    charactersStringArray[i] = charactersStringArray[i].lstrip(" ,\n")
                    charactersStringArray[i] = charactersStringArray[i].rstrip(" ,\n")
            elif characterNameInput == "A":
                charactersStringArray = []
                for person in character_land.charactersPresent:
                    charactersStringArray.append(person.name)
            else:
                charactersStringArray = [characterNameInput]

            for i in range(len(charactersStringArray)):
                if charactersStringArray[i] in CHARACTER_ABBREVIATIONS:
                    charactersStringArray[i] = CHARACTER_ABBREVIATIONS[charactersStringArray[i]]

            characterObjectArray = []
            for characterName in charactersStringArray:
                if characterName == "":
                    continue
 
                if not any(char.name == characterName or (char.name == "Battalion" and str(char.level) == characterName) for char in character_land.charactersPresent):
                    move_input = input(f'Invalid character. "{characterName}" is not present in the selected city. \nPlease enter request separated by a comma and space. \nStart from the beginning with the format [location] [character] [direction] or press x to exit:\n')
                    error = True
                    break
                else:
                    for char in character_land.charactersPresent:
                        if char.name == characterName or (char.name == "Battalion" and str(char.level) == characterName):
                            if char not in characterObjectArray:
                                characterObject = char
                                break

                if characterObject is None:
                    move_input = input("No character selelcted. \nStart from the beginning with the format [location] [character] [direction] or press x to exit:\n")
                    error = True
                    break

                if characterObject.lastTurnTheyMoved == turn:
                    move_input = input("Character has already moved this turn. \nStart from the beginning with the format [location] [character] [direction] or press x to exit:\n")
                    error = True
                    break

                characterObjectArray.append(characterObject)

            if error:
                continue

            if len(characterObjectArray) == 0:
                move_input = input("No characters at this location. \nVerify that you have selected the right location and start from the beginning with the format [location] [character] [direction] or press x to exit:\n")
                continue

            move_row, move_col = MOVEMENT_DIRECTIONS[direction]
            target_row = character_land.row + move_row + 8
            target_col = character_land.column + move_col + 8

            if not (8 <= target_row < (len(world_map)+8) and 8 <= target_col < (len(world_map[0])+8)):
                move_input = input("Cannot move in the specified direction because the target location is outside the map boundaries. \nStart from the beginning with the format [location] [character] [direction] or press x to exit:\n")
                continue

            break

        target_land = world_map[target_row-8][target_col-8]
        target_land.row = target_row-8
        target_land.column = target_col-8

        if target_land.owner == self.name:
            self.move(character_land, characterObjectArray, target_land)
        elif target_land.owner == None:
            self.move(character_land, characterObjectArray, target_land, True)
        else:
            hasAttackerMoved = self.battle(character_land, characterObjectArray, target_land, world_map, kingdoms)
            if hasAttackerMoved == False:
                return input("Attacker Retreating")
    
        for person in characterObjectArray:
            person.lastTurnTheyMoved = turn

    def check_ruler_health(self):
        if self.ruler.health <= 0:
            del self

    def add_land(self, land):
        self.lands.append(land)
        land.city_index = len(self.lands) - 1
        land.owner = self.name

    def remove_land(self, land):
        try:
            self.lands.remove(land)
        except:
            print(f"player" + str(self.playerNumber) + " does not own this land!!")
            self.remove_land(input("Please type in anoter land:\n")) 