import xlwings as xw
from openpyxl.utils import get_column_letter
import random
from termcolor import colored
import openpyxl as px
import os
from battalion import *
from land import *

def clear_screen():
    os.system('cls' if os.name == 'nt' else 'clear')

def center(text, characters=" "):
    width = os.get_terminal_size().columns
    centered_text = characters * ((width - len(text)) // 2) + text + characters * ((width - len(text)) // 2)
    print(centered_text)

# class Battalion:
#     def __init__(self, level=1, name="Battalion", health=1, isRuler=False):
#         self.level = level
#         self.name = name
#         self.health = health
#         self.isRuler = isRuler
#         self.lastTurnTheyMoved = 0

#     def display_health(self):
#         result = ""
#         for _ in range(self.health):
#             result += "💗"
#         return result

# class Land:
#     def __init__(self, name, owner=None, numPeasants=1, housing=4, charactersPresent=None, is_captial=False):
#         self.name = name
#         self.numPeasants = numPeasants
#         self.numAvailablePeasants = numPeasants
#         self.housing = housing
#         self.owner = owner
#         self.is_capital = is_captial
#         if charactersPresent is None:
#             self.charactersPresent = []
#         else:
#             self.charactersPresent = charactersPresent


class Kingdom:
    defeated = False
    gold = 21


    def __init__(self, playerNumber, name, capitalName, ruler, advisor, warrior, color):
        self.playerNumber = playerNumber
        self.name = name
        self.ruler = Battalion(5, ruler, 3, True)
        self.advisor = Battalion(5, advisor, 3)
        self.capital = Land(capitalName, name, 3, 12, [self.ruler, self.advisor], True)
        self.lands = [self.capital]
        self.warrior = warrior
        self.purchase_delay = {}
        self.purchaseWaitTime = {"p": 1, "g": 1, "b": 1, "m": 2, "t": 1, "c": 0, "h": 2, "hp": 0}
        self.color = color
        self.updateVariables()


    def updateVariables(self):
        productionPower = 0
        numCitiesCounter = 0
        for land in self.lands:
            productionPower += land.numPeasants
            numCitiesCounter += 1
        self.production = productionPower
        self.numCities = numCitiesCounter

        if self.purchase_delay:
            for countdown in range(max(self.purchase_delay.keys()), -1, -1):
                if countdown in self.purchase_delay:
                    items = self.purchase_delay[countdown]
                    if countdown == 1:
                        self.addItem(items)
                        del self.purchase_delay[countdown]
                    else:
                        self.purchase_delay[countdown] = self.purchase_delay.get(countdown - 1, [])
                        if not self.purchase_delay[countdown - 1]:
                            del self.purchase_delay[countdown - 1]


    def updateExcel(self):

        wb = xw.Book('World Map.xlsx')
        sht1 = wb.sheets['Sheet1']

        # Loop through each land
        for land in self.lands:
            # Get the row and column of the land in the spreadsheet
            row = land.row + 8
            column = land.column + 8

            # Get the characters present in the land
            character_info = []
            for char in land.charactersPresent:
                if char.name == "Battalion":
                    character_info.append(str(char.level))
                else:
                    character_info.append(char.name)

            # Create a string with the land name and characters
            land_info = land.name + '\n' + ', '.join(character_info)

            # Write the land info to the corresponding cell in the spreadsheet
            cell = sht1.range((row, column))
            cell.value = land_info

        # Save and close the spreadsheet
        wb.save()

    def canMoveStatus(self, character, turn):
        if(character.lastTurnTheyMoved < turn):
            return colored(str(character.level), 'green')
        else:
            return colored(str(character.level),'red')

    def printCharacters(self, charactersPresent, turn):
        result = ""
        for character in charactersPresent:
            if character == self.ruler or character == self.advisor or character.name == "C":
                result += character.name + "<"
            result += "(" + self.canMoveStatus(character, turn) + ") "
        return result
    
    def printPurchaseDelay(self, city_index):
        if city_index not in range(len(self.lands)):
            return ""  # return empty string if city_index is invalid
        if self.purchase_delay:
            output = ""
            for key in self.purchase_delay:
                if isinstance(self.purchase_delay[key][0], list):
                    for arr in self.purchase_delay[key]:
                        if arr[0] == city_index:  # check if the first element is equal to city_index
                            output += f"{'' if arr[1] == 1 else arr[1]}{arr[2].upper()}({self.purchaseWaitTime[arr[2]] - key}/{self.purchaseWaitTime[arr[2]]}) "

                else:
                    if self.purchase_delay[key][0] == city_index:  # check if the first element is equal to city_index
                        output += f"{'' if self.purchase_delay[key][1] == 1 else self.purchase_delay[key][1]}{self.purchase_delay[key][2]}{self.purchaseWaitTime[self.purchase_delay[key][2]] - key}/{self.purchaseWaitTime[self.purchase_delay[key][2]]} "
            return output
        else:
            return ""



    def printCities(self, turn):
        for i in range(len(self.lands)):
            if self.capital == self.lands[i]:
                capitalStar = "*"
            else:
                capitalStar = ""
            print("\n\t" + str(i+1) + " - " + capitalStar + self.lands[i].name + capitalStar + "\n\t\tPeasants: " + str(self.lands[i].numPeasants) + "/" + str(self.lands[i].housing) + "    Idle: " + str(self.lands[i].numAvailablePeasants) + "/" + str(self.lands[i].numPeasants))
            print("\t\tProgress: " + self.printPurchaseDelay(i))
            print("\t\tBattalions: " + self.printCharacters(self.lands[i].charactersPresent, turn))


    def buy(self, item):
        market = {"p": 3, "g": 0, "b": 7, "m": 7, "t": 1, "c": 5, "h": 10, "hp": 10}
        while True:
            if item == "x":
                return
            if item == "shibboleth":
                self.addItem(input("How much money do you want to add? "))
            request = item.split()
            # print(request)
            if len(request) != 3:
                item = input("Invalid input format\nmust be 3 elements long with spaces in between:\n")
                continue

            try:
                city_index = int(request[0]) - 1  # Subtract 1 to convert to 0-based indexing
                quantity = int(request[1])
            except ValueError:
                item = input("Invalid input format\nmust be: 'number, number, letter':\n")
                continue

            item_type = request[2].lower()

            if item_type not in market:
                item = input("Invalid item type\nmust be one of: 'p', 'g', 'b', 'm', 't', 'c', 'h', 'hp' or press x to exit:\n")
                continue

            if self.gold < quantity * market[item_type]:
                item = input("Insufficient gold to purchase the requested quantity of this item:\n")
                continue

            if city_index >= len(self.lands) or city_index < 0:
                item = input("You selected a kingdom that does not exist. \nPlease retry or exit with x:\n")
                continue

            if self.lands[city_index].numAvailablePeasants < quantity:
                item = input("Insufficient peasants to build the requested quantity of this item:\n")
                continue

            break

        # Subtract the cost of the item from the city's gold and subtract one peasant per item built
        self.gold -= quantity * market[item_type]

        # Increase the quantity of the item in the city's inventory
        if item_type == "hp" or item_type == "c" or item_type == "t":
            # print(f"Kingdom index: {city_index + 1}")
            # print(f"Number of items: {quantity}")
            # print(f"Item type: {item_type}")
            # input("Does the above line look ok?")
            self.addItem([city_index, quantity, item_type])
            
            return
        else:
            self.lands[city_index].numAvailablePeasants -= quantity
            self.purchase_delay.setdefault(1, []).append([city_index, quantity, item_type])

        # print(f"Kingdom index: {city_index + 1}")
        # print(f"Number of items: {quantity}")
        # print(f"Item type: {item_type}")
        # print(self.purchase_delay)
        # input("Does the above line look ok?")


    def addItem(self, request):
        if type(request) == str:
            request = request.split()
            request[0] = int(request[0]) - 1 
            request[1] = int(request[1])
        if type(request) == list:
            if type(request[0]) == list:
                # if request is a list of lists
                for array in request:
                    self.addItem(array)
            else:
                # if the request is just a list. 
                city_index = request[0]  # Subtract 1 to convert to 0-based indexing
                quantity = request[1]
                item_type = request[2]


                # CHANGE THIS TO SWITCH SYNTAX
                if item_type == "p":
                    self.lands[city_index].numPeasants += quantity
                    self.lands[city_index].numAvailablePeasants += quantity
                elif item_type == "g":
                    self.gold += quantity
                elif item_type == "b":
                    for _ in range(quantity):
                        self.lands[city_index].charactersPresent.append(Battalion())
                elif item_type == "m":
                    # do something if item_type is "m"
                    print("NOT YET IMPLEMENTED")
                elif item_type == "t":
                    # do something if item_type is "t"
                    print("NOT YET IMPLEMENTED")
                elif item_type == "c":
                    # do something if item_type is "c"
                    for _ in range(quantity):
                        self.lands[city_index].charactersPresent.append(Battalion(3, "C"))
                    return
                elif item_type == "h":
                    # do something if item_type is "h"
                    self.lands[city_index].housing += quantity * 4
                elif item_type == "hp":
                    # do something if item_type is "hp"
                    print("NOT YET IMPLEMENTED")
                else:
                    # handle the case where item_type is not in the products dictionary
                    print(f"There is an error in the request to add an item. {request} is not formatted correctly! Sucks to suck!")

                # input("about to add num available peasants!")
                self.lands[city_index].numAvailablePeasants += quantity

    def battleScreen(self, character_land, characterObjectArray, target_land):
        clear_screen()
        center("-","-")
        center("=","=")
        center(f"   {character_land.owner} Has Attacked {target_land.owner}!!!   ","-")
        center("=","=")
        center("-","-")
        center("-"*len(f"|{character_land.owner}'s army|")," ")
        center(f"|{character_land.owner}'s army|"," ")
        center("-"*len(f"|{character_land.owner}'s army|")," ")
        center(" "," ")
        attackerString = ''
        defenderString = ''
        for i in range(len(characterObjectArray)):
            # attackerString += characterObjectArray[i].name 
            # if i != 0 and i % 6 == 0:
            #     attackerString += "\n"
            if characterObjectArray[i].name != "Battalion":
                attackerString += characterObjectArray[i].name + "<"
                attackerString += "(" + str(characterObjectArray[i].level) + ")"
            # if characterObjectArray[i].name != "Battalion":
                attackerString += " " + characterObjectArray[i].display_health()
                attackerString += " "
            else:
                attackerString += "(" + str(characterObjectArray[i].level) + ")"


        for i in range(len(target_land.charactersPresent)):
            # attackerString += characterObjectArray[i].name 
            # if i != 0 and i % 6 == 0:
            #     defenderString += "\n"
            if target_land.charactersPresent[i].name != "Battalion":
                defenderString += target_land.charactersPresent[i].name + "<"
                defenderString += "(" + str(target_land.charactersPresent[i].level) + ")"
            # if target_land.charactersPresent[i].name != "Battalion":
                defenderString += " " + target_land.charactersPresent[i].display_health()
                defenderString += " "
            else:
                defenderString += "(" + str(target_land.charactersPresent[i].level) + ")"

        center(attackerString," ")
        center("-","-")
        center("-"*len(f"|{target_land.owner}'s army|")," ")
        center(f"|{target_land.owner}'s army|"," ")
        center("-"*len(f"|{target_land.owner}'s army|")," ")
        center(" "," ")
        center(defenderString," ")
        center("","=")

        numbers = [0.14, 0.33, 0.60, 1, 1.66, 3, 7]
        percentage = len(characterObjectArray) / len(target_land.charactersPresent)
        # Find the closest number in the array
        closest_number = min(numbers, key=lambda x: abs(x - percentage))
        # Define unique strings for each case
        output_strings = {
            0.14: "MAX: 1 Attacker vs. 7 Defenders",
            0.33: "MAX: 2 Attackers vs. 6 Defenders",
            0.60: "MAX: 3 Attackers vs. 5 Defenders",
            1: "MAX: 4 Attackers vs. 4 Defenders",
            1.66: "MAX: 5 Attackers vs. 3 Defenders",
            3: "MAX: 6 Attackers vs. 2 Defenders",
            7: "MAX: 7 Attackers vs 1 Defender"
        }
        center(f"|| {output_strings[closest_number]} ||","-")
        center("","=")

        print("    1: Attacker Retreat    3: Input Casualties    5: Something Else")
        print("    2: Defender Retreat    4: Use Artifacts       6: And Something Else")
        return

    def battle(self, character_land, characterObjectArray, target_land, world_map, kingdoms):


        hasAttackerMoved = False
        while True:
            self.battleScreen(character_land, characterObjectArray, target_land)
            option = input("Select an Option: ")
            # if len(target_land.charactersPresent) == 0 | len(characterObjectArray) == 0:
            #     return hasAttackerMoved
            if option == "1":
                return hasAttackerMoved
            elif option == "2":
                directions = {'up': (-1, 0), 'right': (0, 1), 'down': (1, 0), 'left': (0, -1), 'u': (-1, 0), 'r': (0, 1), 'd': (1, 0), 'l': (0, -1)}
                retreatDirection = input("which direction do you want to retreat? ")
                while True:
                    while retreatDirection not in directions:
                        self.battleScreen(character_land, characterObjectArray, target_land)
                        retreatDirection = input("invalid direction. Please type direction or u, d, l, r as an abbreviation ")
                
                    move_row, move_col = directions[retreatDirection]
                    target_row = target_land.row + move_row + 8
                    target_col = target_land.column + move_col + 8

                    if not (8 <= target_row < (len(world_map)+8) and 8 <= target_col < (len(world_map[0])+8)):
                        self.battleScreen(character_land, characterObjectArray, target_land)
                        retreatDirection = input("Cannot move in the specified direction because the target location is outside the map boundaries. \nPlease retype direction:\n")
                        continue

                    new_target_land = world_map[target_row-8][target_col-8]
                    new_target_land.row = target_row-8
                    new_target_land.column = target_col-8

                # make it so that after tey move teir turn ends
                # self.lands[city_index].charactersMoved[character] = True


                    if new_target_land.owner != target_land.owner:
                        self.battleScreen(character_land, characterObjectArray, target_land)
                        retreatDirection = input("You are trying to retreat to a land that you don't own. Please choose another land.\nPlease retype direction:\n")
                        continue


                    # I NEED TO GET THE CITY INDEX HERE
                    targetLandIndexForRetreatingKingdom = kingdoms[new_target_land.owner].lands.index(target_land)
                    targetLandIndexForConqueringKingdom = len(kingdoms[character_land.owner].lands)


                    self.move(target_land, target_land.charactersPresent, new_target_land)


                    kingdoms[new_target_land.owner].remove_land(target_land)



                    self.move(character_land, characterObjectArray, target_land, True)

                    # kingdoms[character_land.owner].updateExcel()
                    # kingdoms[new_target_land.owner].updateExcel()

                    # print(character_land.owner)
                    # print(kingdoms[character_land.owner].purchase_delay)
                    # print(kingdoms[character_land.owner].lands)
                    # print("--------------------------------------------")
                    # print(new_target_land.owner)
                    # print(kingdoms[new_target_land.owner].purchase_delay)
                    # print(kingdoms[new_target_land.owner].lands)
                    # print(target_land.charactersPresent)

                    # input("everything should be updated on the excel")






                    purchase_delay_change = None
                    turns_til_completion = None
                    indexesToDeleteFromNewTargetLandsPurchaseDelay = []
                    
                    for key, value in list(kingdoms[new_target_land.owner].purchase_delay.items()):
                        for i in range(len(value)):
                            print("I'm checking to see if", targetLandIndexForRetreatingKingdom ,"is equivalent to the following city index in the retreating purchase delay: ", value[i][0])
                            input("PURCHASE DELAY IS CHECKED")

                            if value[i][0] == targetLandIndexForRetreatingKingdom:  # Check if the first element matches the desired index
                                input("purchase delay is updated!!!!")
                                purchase_delay_change = value[i] #this should be an array
                                print("purchase_delay_change",purchase_delay_change)
                                turns_til_completion = key if purchase_delay_change[2] == "b" else key - 1
                                print("i:", i, "\nvalue at i:", value[i])
                                indexesToDeleteFromNewTargetLandsPurchaseDelay.append(i)
                    
                            
                            

                            # this only is true if there is something being built in the city that is being conquered
                            if purchase_delay_change != None:
                                print("adding ", purchase_delay_change[1], " of ", purchase_delay_change[2], " to city number ", targetLandIndexForConqueringKingdom, "for the attacking group")
                                input()

                                newEntry = [ targetLandIndexForConqueringKingdom ,purchase_delay_change[1], purchase_delay_change[2]]

                                if turns_til_completion or purchase_delay_change[2] == 'b':
                                    print(newEntry)
                                    input("purchase delay is being altered!!")

                                    self.purchase_delay.setdefault(1, []).append(newEntry)
                                    print(kingdoms[character_land.owner].purchase_delay)
                                    input("update was made to the purchase delay!")
                                elif turns_til_completion == 0 and purchase_delay_change[2] != "b":
                                    self.addItem(newEntry)
                                
                                purchase_delay_change = None
                        
                        for index in indexesToDeleteFromNewTargetLandsPurchaseDelay:
                            value.pop(index)
                            # input("this pops all the things from the purchase delay that are transferred when the land is conquered")

                        # print(character_land.owner)
                        # print(kingdoms[character_land.owner].purchase_delay)
                        # print(kingdoms[character_land.owner].lands)
                        # print("--------------------------------------------")
                        # print(new_target_land.owner)
                        # print(kingdoms[new_target_land.owner].purchase_delay)
                        # print(kingdoms[new_target_land.owner].lands)
                        # input()

                        kingdoms[character_land.owner].updateExcel()
                        kingdoms[new_target_land.owner].updateExcel()

                        hasAttackerMoved = True
                        return hasAttackerMoved
            elif option == "3":
            

                abbreviations = {'g': "Ganondorf", 'z': "Zelda", 'pt': "Palutena", 'n': "Ness", 'b': "Bowser", 'bj': "Bowser Jr.", 'l': "Lucina", 'i': "Ike", 'p': "Peach", 'r': "Rosalina", 'zs': "Zero Suit Samus", 's': "Samus", 'lc': "Lucario", 'gr': "Greninja", 'kd': "King Dedede", 'k': "Kirby"}


                allCharactersString = input(f"please select all of {character_land.owner}'s casualties:\n")

                charactersStringArray = allCharactersString.split(" ")



                for i in range(len(charactersStringArray)):
                    charactersStringArray[i] = charactersStringArray[i].lstrip(" ,\n")
                    charactersStringArray[i] = charactersStringArray[i].rstrip(" ,\n")


                #loop through charactersStringArray for abbreviations here
                for i in range(len(charactersStringArray)):
                    if charactersStringArray[i] in abbreviations:
                        charactersStringArray[i] = abbreviations[charactersStringArray[i]]


                for characterName in charactersStringArray:


                    # Remove unwanted characters from the start and end of each selected character's name

                    print("This is the characters String array after splitting along whitespace", charactersStringArray)

                    if allCharactersString == "x" or allCharactersString == "":
                        break 
                    
                    hasAttackerMoved = True

                    # if characterName == "":
                    #     continue
    
                    if not any(char.name == characterName or (char.name == "Battalion" and str(char.level) == characterName) for char in character_land.charactersPresent):
                        # Character is not present in the array
                        move_input = input(f'Invalid character. "{characterName}" is not present in the attacking city. \nPlease enter request separated by a comma and space:\n')

                        continue
                    else:
                    # Find the character object in the charactersPresent list
                    # characterObject = next((char for char in character_land.charactersPresent if char.name == characterName or (char.name == "Battalion" and str(char.level) == characterName)), None)

                    # Iterate through the character objects
                        print([character.name for character in character_land.charactersPresent])

                        input("above are all the characters present")

                        for i in range(len(character_land.charactersPresent)):
                            char = character_land.charactersPresent[i]
                            if char.name == characterName or (char.name == "Battalion" and str(char.level) == characterName):


                                if (char.decrease_health()):

                                    print("about to pop index", i)
                                    input()
                                    character_land.charactersPresent.pop(i)

                                break

                # STARTING DEFENDER CASUALTIES:

                allCharactersString = input(f"please select all of {target_land.owner}'s casualties:\n")


                charactersStringArray = allCharactersString.split(" ")

                for i in range(len(charactersStringArray)):
                    charactersStringArray[i] = charactersStringArray[i].lstrip(" ,\n")
                    charactersStringArray[i] = charactersStringArray[i].rstrip(" ,\n")


                #loop through charactersStringArray for abbreviations here
                for i in range(len(charactersStringArray)):
                    if charactersStringArray[i] in abbreviations:
                        charactersStringArray[i] = abbreviations[charactersStringArray[i]]


                
                for characterName in charactersStringArray:

                    print("This is the characters String array after splitting along whitespace", charactersStringArray)

                    # Remove unwanted characters from the start and end of each selected character's name

                    if allCharactersString == "x" or allCharactersString == "":
                        break 
                    
                    hasAttackerMoved = True
    
                    if not any(char.name == characterName or (char.name == "Battalion" and str(char.level) == characterName) for char in character_land.charactersPresent):
                        # Character is not present in the array
                        move_input = input(f'Invalid character. "{characterName}" is not present in the attacking city. \nPlease enter request separated by a comma and space:\n')

                        continue
                    else:


                    # Iterate through the character objects
                        print([character.name for character in target_land.charactersPresent])

                        input("above are all the characters present")

                        for i in range(len(target_land.charactersPresent)):
                            char = target_land.charactersPresent[i]
                            if char.name == characterName or (char.name == "Battalion" and str(char.level) == characterName):


                                if (char.decrease_health()):

                                    print("about to pop index", i)
                                    input()
                                    target_land.charactersPresent.pop(i)
                                break


                kingdoms[character_land.owner].updateExcel()
                kingdoms[target_land.owner].updateExcel()


                print(character_land.owner)
                print([character.name for character in character_land.charactersPresent])
                print(kingdoms[character_land.owner].lands)
                print("--------------------------------------------")
                print(target_land.owner)
                print([character.name for character in target_land.charactersPresent])
                print(kingdoms[target_land.owner].lands)
                input()



                # YOU HOULD CHECK THE RULERS HEALTH BEFORE CONTINUING

                
                continue




            elif option == "4":
                pass
            hasAttackerMoved = True

        return hasAttackerMoved

    def move(self, starting_land, characters_to_move, target_land, colonize=False):
        wb = xw.Book('World Map.xlsx')
        sht1 = wb.sheets['Sheet1']

        for character in characters_to_move.copy():

            starting_land.charactersPresent.remove(character)
            target_land.charactersPresent.append(character)


        target_column_letter = get_column_letter(target_land.column + 8)
        target_cell = sht1.range(target_column_letter + str(target_land.row + 8))



        if colonize:
            target_land.owner = self.name
            target_cell.color = self.color
            self.add_land(target_land)


        self.updateExcel()



    def moveRequest(self, world_map, move_input, turn, kingdoms):
        directions = {'up': (-1, 0), 'right': (0, 1), 'down': (1, 0), 'left': (0, -1), 'u': (-1, 0), 'r': (0, 1), 'd': (1, 0), 'l': (0, -1)}
        move_format = ['up', 'right', 'down', 'left', 'u', 'r', 'd', 'l']

        while True:
            error = False

            if move_input == "x":
                return

            move_data = move_input.split()
            if len(move_data) != 3:
                move_input = input("Invalid input format. \nPlease provide three elements separated by spaces with the format [location] [character] [direction] or press x to exit:\n")
                continue

            city_index = int(move_data[0]) - 1
            characterNameInput = move_data[1]
            direction = move_data[2].lower()

            if city_index < 0 or city_index >= len(self.lands):
                move_input = input("Invalid city index. \nPlease select a valid city with the format [location] [character] [direction] or press x to exit:\n")
                continue

            character_land = self.lands[city_index]

            if direction not in move_format:
                move_input = input("Invalid direction. \nPlease choose one of the following: 'up', 'right', 'down', 'left', or use their abbreviations 'u', 'r', 'd', 'l'. \nStart from the beginning with the format [location] [character] [direction] or press x to exit:\n")
                continue

                #---- check to see if multiple are selected ----#

            #dictionary with abbreviations up here
            abbreviations = {'g': "Ganondorf", 'z': "Zelda", 'pt': "Palutena", 'n': "Ness", 'b': "Bowser", 'bj': "Bowser Jr.", 'l': "Lucina", 'i': "Ike", 'p': "Peach", 'r': "Rosalina", 'zs': "Zero Suit Samus", 's': "Samus", 'lc': "Lucario", 'gr': "Greninja", 'kd': "King Dedede", 'k': "Kirby"}

            if characterNameInput == "M":
                allCharactersString = input(f"please select all characters in {character_land.name} you wish to move:\n")
                charactersStringArray = allCharactersString.split(" ")
                # Remove unwanted characters from the start and end of each selected character's name
                for i in range(len(charactersStringArray)):
                    charactersStringArray[i] = charactersStringArray[i].lstrip(" ,\n")
                    charactersStringArray[i] = charactersStringArray[i].rstrip(" ,\n")
            elif characterNameInput == "A":
                charactersStringArray = []
                for person in character_land.charactersPresent:
                    charactersStringArray.append(person.name)
            else:
                charactersStringArray = [characterNameInput]


            #loop through charactersStringArray for abbreviations here
            for i in range(len(charactersStringArray)):
                if charactersStringArray[i] in abbreviations:
                    charactersStringArray[i] = abbreviations[charactersStringArray[i]]


            characterObjectArray = []
            for characterName in charactersStringArray:

                if characterName == "":
                    continue
 
                if not any(char.name == characterName or (char.name == "Battalion" and str(char.level) == characterName) for char in character_land.charactersPresent):
                    # Character is not present in the array
                    move_input = input(f'Invalid character. "{characterName}" is not present in the selected city. \nPlease enter request separated by a comma and space. \nStart from the beginning with the format [location] [character] [direction] or press x to exit:\n')
                    error = True
                    break
                else:
                    # Find the character object in the charactersPresent list
                    # characterObject = next((char for char in character_land.charactersPresent if char.name == characterName or (char.name == "Battalion" and str(char.level) == characterName)), None)

                    # Iterate through the character objects
                    for char in character_land.charactersPresent:
                        if char.name == characterName or (char.name == "Battalion" and str(char.level) == characterName):
                            # Check if the character object is not already in the array
                            if char not in characterObjectArray:
                                # Add the character object to the array
                                # characterObjectArray.append(char)
                                characterObject = char  # Assign the character object to characterObject
                                break  # Exit the loop after finding a suitable character


                if characterObject is None:
                    # Character not found
                    move_input = input("No character selelcted. \nStart from the beginning with the format [location] [character] [direction] or press x to exit:\n")
                    error = True
                    break

        #         Check if th character has already moved
                if characterObject.lastTurnTheyMoved == turn:
                    move_input = input("Character has already moved this turn. \nStart from the beginning with the format [location] [character] [direction] or press x to exit:\n")
                    error = True
                    break

                characterObjectArray.append(characterObject)

            if error:
                continue

            if len(characterObjectArray) == 0:
                move_input = input("No characters at this location. \nVerify that you have selected the right location and start from the beginning with the format [location] [character] [direction] or press x to exit:\n")
                continue

            move_row, move_col = directions[direction]
            target_row = character_land.row + move_row + 8
            target_col = character_land.column + move_col + 8

            # input(target_col)
            # input(target_row)

            if not (8 <= target_row < (len(world_map)+8) and 8 <= target_col < (len(world_map[0])+8)):
                move_input = input("Cannot move in the specified direction because the target location is outside the map boundaries. \nStart from the beginning with the format [location] [character] [direction] or press x to exit:\n")
                continue

            break

        target_land = world_map[target_row-8][target_col-8]
        target_land.row = target_row-8
        target_land.column = target_col-8

        # make it so that after they move their turn ends
        # self.lands[city_index].charactersMoved[character] = True


        if target_land.owner == self.name:
            self.move(character_land, characterObjectArray, target_land)

        elif target_land.owner == None:
            self.move(character_land, characterObjectArray, target_land, True)
            # print("all our lands:", [land.name for land in self.lands])
            input("Colonize Land")

        else:
            #YET TO BE IMPLEMENTED!!!!!!!!!
            input("Battle!")
            hasAttackerMoved = self.battle(character_land, characterObjectArray, target_land, world_map, kingdoms)
            if hasAttackerMoved == False:
                return input("Attacker Retreating")
            # Add an IF statement to see if you win or not
            # YOU NEED TO MOVE THE LOSING ENEMY TEAM OUT FIRST IF YOU LOSE

            # maybe delete this move call? so that when yoiu retreat you can not move
            # self.move(character_land, characterObjectArray, target_land, True)
    
        for person in characterObjectArray:
            person.lastTurnTheyMoved = turn
        # self.lands.append(target_land)
        # input("Move successful. Character has been moved to the target land.")
        # Call add_land function here if needed




    def check_ruler_health(self):
        if self.ruler.health <= 0:
            del self

    def add_land(self, land):
        self.lands.append(land)
        land.city_index = len(self.lands) - 1
        land.owner = self.name

    def remove_land(self, land):
        try:
            self.lands.remove(land)
        except:
            print(f"player" + str(self.playerNumber) + " does not own this land!!")
            self.remove_land(input("Please type in anoter land:\n"))

